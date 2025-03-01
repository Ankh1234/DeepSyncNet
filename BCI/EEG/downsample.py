import openpyxl
import numpy as np

def downsample_eeg(file_path, old_rate=200, new_rate=120):

    # 1) 读取 Excel 文件中的数据
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    worksheet = workbook.active

    # 读取所有行数据（假设每一行是一个时间点，每列是一个通道）
    all_data = []
    for row in worksheet.iter_rows(values_only=True):
        row_vals = row[:30]  # 只取前 30 列（若包含 EOG 通道需自行处理）
        if all(val is not None for val in row_vals):  # 确保行中没有空值
            all_data.append(row_vals)

    old_data = np.array(all_data)  # shape ~ (old_len, 30)

    # 2) 计算新采样后的长度，并使用均匀抽样生成新数据
    ratio = old_rate / new_rate  # 例如: 200 / 120 = 1.6667
    old_len = old_data.shape[0]
    new_len = int(round(old_len * (new_rate / old_rate)))

    new_data_list = []
    for j in range(new_len):
        old_index = int(round(j * ratio))
        if old_index < old_len:
            new_data_list.append(old_data[old_index, :])
        else:
            break

    new_data = np.array(new_data_list)  # 转为 NumPy 数组
    return new_data


def writeExcel(data,datapath):
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    [h, l] = data.shape
    for row in range(1,h+1):
        for col in range(1,l+1):
            outws.cell(row, col).value = data[row-1,col-1]  # 写文件
    outwb.save(datapath)  # 一定要记得保存

for a in range(1,30):

    for b in range(2, 7, 2):
        data_session_extracted = []
        data_path = r'/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/raw_data/MA/sub0'+str(a)+'_session0' + str(
            b) + '_data.xlsx'

        data = downsample_eeg(data_path)
        print(data.shape)
        writeExcel(data , '/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/downsampled_data/MA/sub0' + str(a) + '_session0' + str(b) + '_downsampled_data.xlsx')
