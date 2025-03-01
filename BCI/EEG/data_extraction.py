import sys
import numpy as np
import pandas as pd
import openpyxl
from scipy.interpolate import griddata
import matplotlib.pyplot as plt


def extract1(file, index=0):
    # 加载工作簿
    workbook = openpyxl.load_workbook(file, data_only=True)

    # 获取指定索引的工作表
    sheetnames = workbook.sheetnames
    worksheet = workbook[sheetnames[index]]

    # 获取所有行的数据
    all_data = []
    for row in worksheet.iter_rows(values_only=True):
        if any(val is not None for val in row):  # 跳过全空行
            all_data.append(row)

    # 转换为 NumPy 数组
    cc = np.array(all_data)
    return cc


for a in range(1,30):

    for b in range(2, 7, 2):
        data_session_extracted = []
        data_path = r'/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/downsampled_data/MA/sub0'+str(a)+'_session0' + str(
            b) + '_downsampled_data.xlsx'
        events_start_path = r'/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/label_Hz/MA/sub0'+str(a)+'_session0' + str(
            b) + '_label_Hz.xlsx'
        label_path = r'/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/label/MA/sub0'+str(a)+'_session0' + str(b) + '_label.xlsx'

        data = extract1(data_path)
        label = extract1(label_path)
        # data = data.T
        events = extract1(events_start_path)
        events = events.astype(int)
        label = label.astype(int)

        arr_3D_session = []
        label_s = []
        num_event = 0

        for i in events[0]:
            arr_time_window = []
            if label[num_event] == 16:
                num_event += 1
                for l in range(-2, 8):
                    arr = data[i + l * 120:i + l * 120 + 360]
                    arr_unit = []
                    for c in range(360):
                        arr_3D = np.full((16, 16), np.nan)

                        # 定义插值范围
                        interpolation_ranges = {
                            0: [6, 11 - 1],
                            1: [5, 12 - 1],
                            2: [4, 13 - 1],
                            3: [2, 15 - 1],
                            4: [2, 15 - 1],
                            5: [2, 15 - 1],
                            6: [2, 15 - 1],
                            7: [1, 16 - 1],
                            8: [2, 15 - 1],
                            9: [2, 15 - 1],
                            10: [2, 15 - 1],
                            11: [2, 15 - 1],
                            12: [4, 13 - 1],
                            13: [6, 11 - 1],
                            14: [6, 11 - 1],
                            15: [7, 10 - 1]
                        }

                        # 遍历矩阵行，设置待插值区域为 0，其他区域保持为 NaN
                        for row in range(16):
                            if row in interpolation_ranges:
                                start, end = interpolation_ranges[row]
                                arr_3D[row, start:end + 1] = 0  # 将待插值区域设置为 0
                            else:
                                arr_3D[row, :] = np.nan  # 非插值区域保持为 NaN

                            # 插入通道数据
                        arr_3D[4 - 1, 3 - 1] = arr[c, 0]  # F7
                        arr_3D[3 - 1, 6 - 1] = arr[c, 1]  # AFF5h
                        arr_3D[4 - 1, 6 - 1] = arr[c, 2]  # F3
                        arr_3D[1 - 1, 8 - 1] = arr[c, 3]  # AFp1
                        arr_3D[1 - 1, 10 - 1] = arr[c, 4]  # AFp2
                        arr_3D[3 - 1, 12 - 1] = arr[c, 5]  # AFF6h
                        arr_3D[4 - 1, 12 - 1] = arr[c, 6]  # F4
                        arr_3D[4 - 1, 15 - 1] = arr[c, 7]  # F8
                        arr_3D[3 - 1, 8 - 1] = arr[c, 8]  # AFF1h
                        arr_3D[3 - 1, 10 - 1] = arr[c, 9]  # AFF2h
                        arr_3D[8 - 1, 9 - 1] = arr[c, 10]  # Cz
                        arr_3D[12 - 1, 9 - 1] = arr[c, 11]  # Pz
                        arr_3D[7 - 1, 4 - 1] = arr[c, 12]  # FCC5h
                        arr_3D[7 - 1, 6 - 1] = arr[c, 13]  # FCC3h
                        arr_3D[9 - 1, 4 - 1] = arr[c, 14]  # CCP5h
                        arr_3D[9 - 1, 6 - 1] = arr[c, 15]  # CCP3h
                        arr_3D[8 - 1, 2 - 1] = arr[c, 16]  # T7
                        arr_3D[12 - 1, 3 - 1] = arr[c, 17]  # P7
                        arr_3D[12 - 1, 6 - 1] = arr[c, 18]  # P3
                        arr_3D[13 - 1, 8 - 1] = arr[c, 19]  # PPO1h
                        arr_3D[15 - 1, 8 - 1] = arr[c, 20]  # POO1
                        arr_3D[15 - 1, 10 - 1] = arr[c, 21]  # POO2
                        arr_3D[13 - 1, 10 - 1] = arr[c, 22]  # PPO2h
                        arr_3D[12 - 1, 12 - 1] = arr[c, 23]  # P4
                        arr_3D[7 - 1, 12 - 1] = arr[c, 24]  # FCC4h
                        arr_3D[7 - 1, 14 - 1] = arr[c, 25]  # FCC6h
                        arr_3D[9 - 1, 12 - 1] = arr[c, 26]  # CCP4h
                        arr_3D[9 - 1, 14 - 1] = arr[c, 27]  # CCP6h
                        arr_3D[12 - 1, 15 - 1] = arr[c, 28]  # P8
                        arr_3D[8 - 1, 16 - 1] = arr[c, 29]  # T8


                        # 获取已知点的坐标和值（通过遍历矩阵）
                        known_points = []
                        known_values = []

                        for row in range(16):
                            for col in range(16):
                                if not np.isnan(arr_3D[row, col]) and arr_3D[row, col] != 0:  # 确认值是已知点（非 NaN 且非 0）
                                    known_points.append((row, col))
                                    known_values.append(arr_3D[row, col])

                        # 获取插值区域中的目标点
                        target_points = []
                        for row in range(16):
                            if row in interpolation_ranges:
                                start, end = interpolation_ranges[row]
                                for col in range(start, end + 1):
                                    # 只添加待插值的点（值为 0 的点）
                                    if arr_3D[row, col] == 0:
                                        target_points.append((row, col))

                        # 使用 linear 方法对插值区域进行插值
                        interpolated_values_linear = griddata(
                            points=known_points,  # 已知点坐标
                            values=known_values,  # 已知点的值
                            xi=target_points,  # 插值目标点
                            method='cubic'  # 三次插值
                        )

                        # 使用 nearest 方法对 NaN 边界点进行补充插值
                        interpolated_values_nearest = griddata(
                            points=known_points,  # 已知点坐标
                            values=known_values,  # 已知点的值
                            xi=target_points,  # 插值目标点
                            method='nearest'  # 最近邻插值
                        )

                        # 合并两种插值结果（优先使用 linear 的结果）
                        final_interpolated_values = []
                        for d, val in enumerate(interpolated_values_linear):
                            if np.isnan(val):  # 如果 linear 插值为 NaN，用 nearest 的值补充
                                final_interpolated_values.append(interpolated_values_nearest[d])
                            else:
                                final_interpolated_values.append(val)

                        # 将插值结果覆盖到矩阵中
                        for idx, (row, col) in enumerate(target_points):
                            arr_3D[row, col] = final_interpolated_values[idx]

                        for row in range(16):
                            for col in range(16):
                                if np.isnan(arr_3D[row, col]):  #将空白部分填充为0
                                    arr_3D[row, col] = 0


                        # 可视化插值结果
                        # plt.figure(figsize=(8, 8))
                        # plt.title("Interpolated Matrix")
                        # plt.imshow(arr_3D, cmap='viridis', origin='upper')
                        # plt.colorbar(label='Interpolated Values')
                        # plt.show()

                        arr_3D = arr_3D.tolist()
                        arr_unit.append(arr_3D)
                    label_s.append(16)
                    arr_time_window.append(arr_unit)
            elif label[num_event] == 32:
                num_event += 1
                for l in range(-2, 8):
                    arr = data[i + l * 120:i + l * 120 + 360]
                    arr_unit = []
                    for c in range(360):
                        arr_3D = np.full((16, 16), np.nan)

                        # 定义插值范围
                        interpolation_ranges = {
                            0: [6, 11 - 1],
                            1: [5, 12 - 1],
                            2: [4, 13 - 1],
                            3: [2, 15 - 1],
                            4: [2, 15 - 1],
                            5: [2, 15 - 1],
                            6: [2, 15 - 1],
                            7: [1, 16 - 1],
                            8: [2, 15 - 1],
                            9: [2, 15 - 1],
                            10: [2, 15 - 1],
                            11: [2, 15 - 1],
                            12: [4, 13 - 1],
                            13: [6, 11 - 1],
                            14: [6, 11 - 1],
                            15: [7, 10 - 1]
                        }

                        # 遍历矩阵行，设置待插值区域为 0，其他区域保持为 NaN
                        for row in range(16):
                            if row in interpolation_ranges:
                                start, end = interpolation_ranges[row]
                                arr_3D[row, start:end + 1] = 0  # 将待插值区域设置为 0
                            else:
                                arr_3D[row, :] = np.nan  # 非插值区域保持为 NaN

                        # 插入通道数据
                        arr_3D[4 - 1, 3 - 1] = arr[c, 0]  # F7
                        arr_3D[3 - 1, 6 - 1] = arr[c, 1]  # AFF5h
                        arr_3D[4 - 1, 6 - 1] = arr[c, 2]  # F3
                        arr_3D[1 - 1, 8 - 1] = arr[c, 3]  # AFp1
                        arr_3D[1 - 1, 10 - 1] = arr[c, 4]  # AFp2
                        arr_3D[3 - 1, 12 - 1] = arr[c, 5]  # AFF6h
                        arr_3D[4 - 1, 12 - 1] = arr[c, 6]  # F4
                        arr_3D[4 - 1, 15 - 1] = arr[c, 7]  # F8
                        arr_3D[3 - 1, 8 - 1] = arr[c, 8]  # AFF1h
                        arr_3D[3 - 1, 10 - 1] = arr[c, 9]  # AFF2h
                        arr_3D[8 - 1, 9 - 1] = arr[c, 10]  # Cz
                        arr_3D[12 - 1, 9 - 1] = arr[c, 11]  # Pz
                        arr_3D[7 - 1, 4 - 1] = arr[c, 12]  # FCC5h
                        arr_3D[7 - 1, 6 - 1] = arr[c, 13]  # FCC3h
                        arr_3D[9 - 1, 4 - 1] = arr[c, 14]  # CCP5h
                        arr_3D[9 - 1, 6 - 1] = arr[c, 15]  # CCP3h
                        arr_3D[8 - 1, 2 - 1] = arr[c, 16]  # T7
                        arr_3D[12 - 1, 3 - 1] = arr[c, 17]  # P7
                        arr_3D[12 - 1, 6 - 1] = arr[c, 18]  # P3
                        arr_3D[13 - 1, 8 - 1] = arr[c, 19]  # PPO1h
                        arr_3D[15 - 1, 8 - 1] = arr[c, 20]  # POO1
                        arr_3D[15 - 1, 10 - 1] = arr[c, 21]  # POO2
                        arr_3D[13 - 1, 10 - 1] = arr[c, 22]  # PPO2h
                        arr_3D[12 - 1, 12 - 1] = arr[c, 23]  # P4
                        arr_3D[7 - 1, 12 - 1] = arr[c, 24]  # FCC4h
                        arr_3D[7 - 1, 14 - 1] = arr[c, 25]  # FCC6h
                        arr_3D[9 - 1, 12 - 1] = arr[c, 26]  # CCP4h
                        arr_3D[9 - 1, 14 - 1] = arr[c, 27]  # CCP6h
                        arr_3D[12 - 1, 15 - 1] = arr[c, 28]  # P8
                        arr_3D[8 - 1, 16 - 1] = arr[c, 29]  # T8

                        # 获取已知点的坐标和值（通过遍历矩阵）
                        known_points = []
                        known_values = []

                        for row in range(16):
                            for col in range(16):
                                if not np.isnan(arr_3D[row, col]) and arr_3D[row, col] != 0:  # 确认值是已知点（非 NaN 且非 0）
                                    known_points.append((row, col))
                                    known_values.append(arr_3D[row, col])

                        # 获取插值区域中的目标点
                        target_points = []
                        for row in range(16):
                            if row in interpolation_ranges:
                                start, end = interpolation_ranges[row]
                                for col in range(start, end + 1):
                                    # 只添加待插值的点（值为 0 的点）
                                    if arr_3D[row, col] == 0:
                                        target_points.append((row, col))

                        # 使用 linear 方法对插值区域进行插值
                        interpolated_values_linear = griddata(
                            points=known_points,  # 已知点坐标
                            values=known_values,  # 已知点的值
                            xi=target_points,  # 插值目标点
                            method='cubic'  # 三次插值
                        )

                        # 使用 nearest 方法对 NaN 边界点进行补充插值
                        interpolated_values_nearest = griddata(
                            points=known_points,  # 已知点坐标
                            values=known_values,  # 已知点的值
                            xi=target_points,  # 插值目标点
                            method='nearest'  # 最近邻插值
                        )

                        # 合并两种插值结果（优先使用 linear 的结果）
                        final_interpolated_values = []
                        for d, val in enumerate(interpolated_values_linear):
                            if np.isnan(val):  # 如果 linear 插值为 NaN，用 nearest 的值补充
                                final_interpolated_values.append(interpolated_values_nearest[d])
                            else:
                                final_interpolated_values.append(val)

                        # 将插值结果覆盖到矩阵中
                        for idx, (row, col) in enumerate(target_points):
                            arr_3D[row, col] = final_interpolated_values[idx]

                        for row in range(16):
                            for col in range(16):
                                if np.isnan(arr_3D[row, col]):  #将空白部分填充为0
                                    arr_3D[row, col] = 0

                        # 可视化插值结果
                        #plt.figure(figsize=(8, 8))
                        #plt.title("Interpolated Matrix")
                        #plt.imshow(arr_3D, cmap='viridis', origin='upper')
                        #plt.colorbar(label='Interpolated Values')
                        #plt.show()

                        arr_3D = arr_3D.tolist()
                        arr_unit.append(arr_3D)
                    label_s.append(32)
                    arr_time_window.append(arr_unit)
            else:
                print("Wrong lable")
                sys.exit()

            # arr_time_window = np.array(arr_time_window)
            # print(arr_time_window.shape)
            # arr_time_window = arr_time_window.tolist()
            arr_3D_session.append(arr_time_window)
        data_extracted = np.array(arr_3D_session)
        print(data_extracted.shape)
        data_extracted = data_extracted.reshape(200,-1)
        print(data_extracted.shape)
        if b == 2:
            data_session2_extracted = data_extracted
            label_session2_extracted = label_s
        elif b == 4:
            data_session4_extracted = data_extracted
            label_session4_extracted = label_s
        else:
            data_session6_extracted = data_extracted
            label_session6_extracted = label_s
    data_sub_extracted = np.concatenate((data_session2_extracted, data_session4_extracted, data_session6_extracted),
                                        axis=0)
    label_sub_extracted = np.concatenate((label_session2_extracted, label_session4_extracted, label_session6_extracted),
                                        axis=0)
    print(data_sub_extracted.shape)
    print(label_sub_extracted.shape)
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/extracted_data/MA/sub0'+str(a)+'_extracted_data.csv', data_sub_extracted, delimiter=',')
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/extracted_label/MA/sub0' + str(a) + '_extracted_label.csv', label_sub_extracted, delimiter='')


