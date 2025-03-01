import sys

import numpy as np
import pandas as pd
import openpyxl

import numpy as np
import pandas as pd

from scipy.interpolate import griddata
import matplotlib.pyplot as plt


def extract1(file, index=0):
    # 加载工作簿
    workbook = openpyxl.load_workbook(file, data_only=True)

    # 获取指定索引的工作表
    sheetnames = workbook.sheetnames
    worksheet = workbook[sheetnames[index]]

    # 获取所有行的数据
    all_data = []
    for row in worksheet.iter_rows(values_only=True):
        if any(val is not None for val in row):  # 跳过全空行
            all_data.append(row)

    # 转换为 NumPy 数组
    cc = np.array(all_data)
    return cc


for a in range(1,30):
    arr1 = []
    arr2 = []
    arr3 = []
    arr4 = []
    arr5 = []
    arr6 = []
    arr7 = []
    arr8 = []
    arr9 = []
    arr10 = []

    label1 = []
    label2 = []
    label3 = []
    label4 = []
    label5 = []
    label6 = []
    label7 = []
    label8 = []
    label9 = []
    label10 = []
    for b in range(2, 7, 2):
        data_path = r'/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/downsampled_data/MA/sub0'+str(a)+'_session0' + str(
            b) + '_downsampled_data.xlsx'
        events_start_path = r'/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/label_Hz/MA/sub0'+str(a)+'_session0' + str(
            b) + '_label_Hz.xlsx'
        label_path = r'/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/label/MA/sub0'+str(a)+'_session0' + str(b) + '_label.xlsx'

        data = extract1(data_path)
        label = extract1(label_path)
        # data = data.T
        events = extract1(events_start_path)
        events = events.astype(int)
        label = label.astype(int)
        # print(events.shape)
        print(data.shape)
        # print(events)
        # print(events[0][0])
        num_event = 0
        for i in events[0]:
            if label[num_event] == 16:
                num_event += 1
                for c in range(-2, 8):
                    arr = data[i + c * 120:i + c * 120 + 360]
                    arr_3D_event = []
                    for d in range(360):
                        arr_3D = np.full((16, 16), np.nan)

                        # 定义插值范围
                        interpolation_ranges = {
                            0: [6, 11 - 1],
                            1: [5, 12 - 1],
                            2: [4, 13 - 1],
                            3: [2, 15 - 1],
                            4: [2, 15 - 1],
                            5: [2, 15 - 1],
                            6: [2, 15 - 1],
                            7: [1, 16 - 1],
                            8: [2, 15 - 1],
                            9: [2, 15 - 1],
                            10: [2, 15 - 1],
                            11: [2, 15 - 1],
                            12: [4, 13 - 1],
                            13: [6, 11 - 1],
                            14: [6, 11 - 1],
                            15: [7, 10 - 1]
                        }

                        # 遍历矩阵行，设置待插值区域为 0，其他区域保持为 NaN
                        for row in range(16):
                            if row in interpolation_ranges:
                                start, end = interpolation_ranges[row]
                                arr_3D[row, start:end + 1] = 0  # 将待插值区域设置为 0
                            else:
                                arr_3D[row, :] = np.nan  # 非插值区域保持为 NaN

                        # 插入通道数据
                        arr_3D[4 - 1, 3 - 1] = arr[d, 0]  # F7
                        arr_3D[3 - 1, 6 - 1] = arr[d, 1]  # AFF5h
                        arr_3D[4 - 1, 6 - 1] = arr[d, 2]  # F3
                        arr_3D[1 - 1, 8 - 1] = arr[d, 3]  # AFp1
                        arr_3D[1 - 1, 10 - 1] = arr[d, 4]  # AFp2
                        arr_3D[3 - 1, 12 - 1] = arr[d, 5]  # AFF6h
                        arr_3D[4 - 1, 12 - 1] = arr[d, 6]  # F4
                        arr_3D[4 - 1, 15 - 1] = arr[d, 7]  # F8
                        arr_3D[3 - 1, 8 - 1] = arr[d, 8]  # AFF1h
                        arr_3D[3 - 1, 10 - 1] = arr[d, 9]  # AFF2h
                        arr_3D[8 - 1, 9 - 1] = arr[d, 10]  # Cz
                        arr_3D[12 - 1, 9 - 1] = arr[d, 11]  # Pz
                        arr_3D[7 - 1, 4 - 1] = arr[d, 12]  # FCC5h
                        arr_3D[7 - 1, 6 - 1] = arr[d, 13]  # FCC3h
                        arr_3D[9 - 1, 4 - 1] = arr[d, 14]  # CCP5h
                        arr_3D[9 - 1, 6 - 1] = arr[d, 15]  # CCP3h
                        arr_3D[8 - 1, 2 - 1] = arr[d, 16]  # T7
                        arr_3D[12 - 1, 3 - 1] = arr[d, 17]  # P7
                        arr_3D[12 - 1, 6 - 1] = arr[d, 18]  # P3
                        arr_3D[13 - 1, 8 - 1] = arr[d, 19]  # PPO1h
                        arr_3D[15 - 1, 8 - 1] = arr[d, 20]  # POO1
                        arr_3D[15 - 1, 10 - 1] = arr[d, 21]  # POO2
                        arr_3D[13 - 1, 10 - 1] = arr[d, 22]  # PPO2h
                        arr_3D[12 - 1, 12 - 1] = arr[d, 23]  # P4
                        arr_3D[7 - 1, 12 - 1] = arr[d, 24]  # FCC4h
                        arr_3D[7 - 1, 14 - 1] = arr[d, 25]  # FCC6h
                        arr_3D[9 - 1, 12 - 1] = arr[d, 26]  # CCP4h
                        arr_3D[9 - 1, 14 - 1] = arr[d, 27]  # CCP6h
                        arr_3D[12 - 1, 15 - 1] = arr[d, 28]  # P8
                        arr_3D[8 - 1, 16 - 1] = arr[d, 29]  # T8

                        # 获取已知点的坐标和值（通过遍历矩阵）
                        known_points = []
                        known_values = []

                        for row in range(16):
                            for col in range(16):
                                if not np.isnan(arr_3D[row, col]) and arr_3D[row, col] != 0:  # 确认值是已知点（非 NaN 且非 0）
                                    known_points.append((row, col))
                                    known_values.append(arr_3D[row, col])

                        # 获取插值区域中的目标点
                        target_points = []
                        for row in range(16):
                            if row in interpolation_ranges:
                                start, end = interpolation_ranges[row]
                                for col in range(start, end + 1):
                                    # 只添加待插值的点（值为 0 的点）
                                    if arr_3D[row, col] == 0:
                                        target_points.append((row, col))

                        # 使用 linear 方法对插值区域进行插值
                        interpolated_values_linear = griddata(
                            points=known_points,  # 已知点坐标
                            values=known_values,  # 已知点的值
                            xi=target_points,  # 插值目标点
                            method='linear'  # 线性插值
                        )

                        # 使用 nearest 方法对 NaN 边界点进行补充插值
                        interpolated_values_nearest = griddata(
                            points=known_points,  # 已知点坐标
                            values=known_values,  # 已知点的值
                            xi=target_points,  # 插值目标点
                            method='nearest'  # 最近邻插值
                        )

                        # 合并两种插值结果（优先使用 linear 的结果）
                        final_interpolated_values = []
                        for e, val in enumerate(interpolated_values_linear):
                            if np.isnan(val):  # 如果 linear 插值为 NaN，用 nearest 的值补充
                                final_interpolated_values.append(interpolated_values_nearest[e])
                            else:
                                final_interpolated_values.append(val)

                        # 将插值结果覆盖到矩阵中
                        for idx, (row, col) in enumerate(target_points):
                            arr_3D[row, col] = final_interpolated_values[idx]

                        for row in range(16):
                            for col in range(16):
                                if np.isnan(arr_3D[row, col]):  #将空白部分填充为0
                                    arr_3D[row, col] = 0

                        arr_3D = arr_3D.tolist()
                        arr_3D_event.append(arr_3D)

                    if c == -2:
                        label1.append(16)
                        arr1.append(arr_3D_event)
                    elif c == -1:
                        label2.append(16)
                        arr2.append(arr_3D_event)
                    elif c == 0:
                        label3.append(16)
                        arr3.append(arr_3D_event)
                    elif c == 1:
                        label4.append(16)
                        arr4.append(arr_3D_event)
                    elif c == 2:
                        label5.append(16)
                        arr5.append(arr_3D_event)
                    elif c == 3:
                        label6.append(16)
                        arr6.append(arr_3D_event)
                    elif c == 4:
                        label7.append(16)
                        arr7.append(arr_3D_event)
                    elif c == 5:
                        label8.append(16)
                        arr8.append(arr_3D_event)
                    elif c == 6:
                        label9.append(16)
                        arr9.append(arr_3D_event)
                    elif c == 7:
                        label10.append(16)
                        arr10.append(arr_3D_event)
                    else:
                        print("Wrong lable")
                        sys.exit()

            elif label[num_event] == 32:
                num_event += 1
                for c in range(-2, 8):
                    arr = data[i + c * 120:i + c * 120 + 360]
                    arr_3D_event = []
                    for d in range(360):
                        arr_3D = np.full((16, 16), np.nan)

                        # 定义插值范围
                        interpolation_ranges = {
                            0: [6, 11 - 1],
                            1: [5, 12 - 1],
                            2: [4, 13 - 1],
                            3: [2, 15 - 1],
                            4: [2, 15 - 1],
                            5: [2, 15 - 1],
                            6: [2, 15 - 1],
                            7: [1, 16 - 1],
                            8: [2, 15 - 1],
                            9: [2, 15 - 1],
                            10: [2, 15 - 1],
                            11: [2, 15 - 1],
                            12: [4, 13 - 1],
                            13: [6, 11 - 1],
                            14: [6, 11 - 1],
                            15: [7, 10 - 1]
                        }

                        # 遍历矩阵行，设置待插值区域为 0，其他区域保持为 NaN
                        for row in range(16):
                            if row in interpolation_ranges:
                                start, end = interpolation_ranges[row]
                                arr_3D[row, start:end + 1] = 0  # 将待插值区域设置为 0
                            else:
                                arr_3D[row, :] = np.nan  # 非插值区域保持为 NaN

                        # 插入通道数据
                        arr_3D[4 - 1, 3 - 1] = arr[d, 0]  # F7
                        arr_3D[3 - 1, 6 - 1] = arr[d, 1]  # AFF5h
                        arr_3D[4 - 1, 6 - 1] = arr[d, 2]  # F3
                        arr_3D[1 - 1, 8 - 1] = arr[d, 3]  # AFp1
                        arr_3D[1 - 1, 10 - 1] = arr[d, 4]  # AFp2
                        arr_3D[3 - 1, 12 - 1] = arr[d, 5]  # AFF6h
                        arr_3D[4 - 1, 12 - 1] = arr[d, 6]  # F4
                        arr_3D[4 - 1, 15 - 1] = arr[d, 7]  # F8
                        arr_3D[3 - 1, 8 - 1] = arr[d, 8]  # AFF1h
                        arr_3D[3 - 1, 10 - 1] = arr[d, 9]  # AFF2h
                        arr_3D[8 - 1, 9 - 1] = arr[d, 10]  # Cz
                        arr_3D[12 - 1, 9 - 1] = arr[d, 11]  # Pz
                        arr_3D[7 - 1, 4 - 1] = arr[d, 12]  # FCC5h
                        arr_3D[7 - 1, 6 - 1] = arr[d, 13]  # FCC3h
                        arr_3D[9 - 1, 4 - 1] = arr[d, 14]  # CCP5h
                        arr_3D[9 - 1, 6 - 1] = arr[d, 15]  # CCP3h
                        arr_3D[8 - 1, 2 - 1] = arr[d, 16]  # T7
                        arr_3D[12 - 1, 3 - 1] = arr[d, 17]  # P7
                        arr_3D[12 - 1, 6 - 1] = arr[d, 18]  # P3
                        arr_3D[13 - 1, 8 - 1] = arr[d, 19]  # PPO1h
                        arr_3D[15 - 1, 8 - 1] = arr[d, 20]  # POO1
                        arr_3D[15 - 1, 10 - 1] = arr[d, 21]  # POO2
                        arr_3D[13 - 1, 10 - 1] = arr[d, 22]  # PPO2h
                        arr_3D[12 - 1, 12 - 1] = arr[d, 23]  # P4
                        arr_3D[7 - 1, 12 - 1] = arr[d, 24]  # FCC4h
                        arr_3D[7 - 1, 14 - 1] = arr[d, 25]  # FCC6h
                        arr_3D[9 - 1, 12 - 1] = arr[d, 26]  # CCP4h
                        arr_3D[9 - 1, 14 - 1] = arr[d, 27]  # CCP6h
                        arr_3D[12 - 1, 15 - 1] = arr[d, 28]  # P8
                        arr_3D[8 - 1, 16 - 1] = arr[d, 29]  # T8

                        # 获取已知点的坐标和值（通过遍历矩阵）
                        known_points = []
                        known_values = []

                        for row in range(16):
                            for col in range(16):
                                if not np.isnan(arr_3D[row, col]) and arr_3D[row, col] != 0:  # 确认值是已知点（非 NaN 且非 0）
                                    known_points.append((row, col))
                                    known_values.append(arr_3D[row, col])

                        # 获取插值区域中的目标点
                        target_points = []
                        for row in range(16):
                            if row in interpolation_ranges:
                                start, end = interpolation_ranges[row]
                                for col in range(start, end + 1):
                                    # 只添加待插值的点（值为 0 的点）
                                    if arr_3D[row, col] == 0:
                                        target_points.append((row, col))

                        # 使用 linear 方法对插值区域进行插值
                        interpolated_values_linear = griddata(
                            points=known_points,  # 已知点坐标
                            values=known_values,  # 已知点的值
                            xi=target_points,  # 插值目标点
                            method='cubic'  # 三次插值
                        )

                        # 使用 nearest 方法对 NaN 边界点进行补充插值
                        interpolated_values_nearest = griddata(
                            points=known_points,  # 已知点坐标
                            values=known_values,  # 已知点的值
                            xi=target_points,  # 插值目标点
                            method='nearest'  # 最近邻插值
                        )

                        # 合并两种插值结果（优先使用 linear 的结果）
                        final_interpolated_values = []
                        for e, val in enumerate(interpolated_values_linear):
                            if np.isnan(val):  # 如果 linear 插值为 NaN，用 nearest 的值补充
                                final_interpolated_values.append(interpolated_values_nearest[e])
                            else:
                                final_interpolated_values.append(val)

                        # 将插值结果覆盖到矩阵中
                        for idx, (row, col) in enumerate(target_points):
                            arr_3D[row, col] = final_interpolated_values[idx]

                        for row in range(16):
                            for col in range(16):
                                if np.isnan(arr_3D[row, col]):  #将空白部分填充为0
                                    arr_3D[row, col] = 0

                        arr_3D = arr_3D.tolist()
                        arr_3D_event.append(arr_3D)

                    if c == -2:
                        label1.append(32)
                        arr1.append(arr_3D_event)
                    elif c == -1:
                        label2.append(32)
                        arr2.append(arr_3D_event)
                    elif c == 0:
                        label3.append(32)
                        arr3.append(arr_3D_event)
                    elif c == 1:
                        label4.append(32)
                        arr4.append(arr_3D_event)
                    elif c == 2:
                        label5.append(32)
                        arr5.append(arr_3D_event)
                    elif c == 3:
                        label6.append(32)
                        arr6.append(arr_3D_event)
                    elif c == 4:
                        label7.append(32)
                        arr7.append(arr_3D_event)
                    elif c == 5:
                        label8.append(32)
                        arr8.append(arr_3D_event)
                    elif c == 6:
                        label9.append(32)
                        arr9.append(arr_3D_event)
                    elif c == 7:
                        label10.append(32)
                        arr10.append(arr_3D_event)
                    else:
                        print("Wrong lable")
                        sys.exit()
            else:
                print("Wrong lable")
                sys.exit()

    arr1 = np.array(arr1)
    label1 = np.array(label1)
    label1.T
    label1.astype(int)
    arr1 = arr1.reshape(60, -1)

    arr2 = np.array(arr2)
    label2 = np.array(label2)
    label2.T
    label2.astype(int)
    arr2 = arr2.reshape(60, -1)

    arr3 = np.array(arr3)
    label3 = np.array(label3)
    label3.T
    label3.astype(int)
    arr3 = arr3.reshape(60, -1)

    arr4 = np.array(arr4)
    label4 = np.array(label4)
    label4.T
    label4.astype(int)
    arr4 = arr4.reshape(60, -1)

    arr5 = np.array(arr5)
    label5 = np.array(label5)
    label5.T
    label5.astype(int)
    arr5 = arr5.reshape(60, -1)

    arr6 = np.array(arr6)
    label6 = np.array(label6)
    label6.T
    label6.astype(int)
    arr6 = arr6.reshape(60, -1)

    arr7 = np.array(arr7)
    label7 = np.array(label7)
    label7.T
    label7.astype(int)
    arr7 = arr7.reshape(60, -1)

    arr8 = np.array(arr8)
    label8 = np.array(label8)
    label8.T
    label8.astype(int)
    arr8 = arr8.reshape(60, -1)

    arr9 = np.array(arr9)
    label9 = np.array(label9)
    label9.T
    label9.astype(int)
    arr9 = arr9.reshape(60, -1)

    arr10 = np.array(arr10)
    label10 = np.array(label10)
    label10.T
    label10.astype(int)
    arr10 = arr10.reshape(60, -1)

    print(arr1.shape)
    print(label1.shape)


    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_data/MA/sub0' + str(a) + '_test_data_00.csv', arr1, delimiter=',')
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_label/MA/sub0' + str(a) + '_test_label_00.csv', label1, delimiter='')

    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_data/MA/sub0' + str(a) + '_test_data_01.csv',arr2, delimiter=',')
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_label/MA/sub0' + str(a) + '_test_label_01.csv',label2, delimiter='')

    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_data/MA/sub0' + str(a) + '_test_data_02.csv',arr3, delimiter=',')
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_label/MA/sub0' + str(a) + '_test_label_02.csv',label3, delimiter='')

    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_data/MA/sub0' + str(a) + '_test_data_03.csv',arr4, delimiter=',')
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_label/MA/sub0' + str(a) + '_test_label_03.csv',label4, delimiter='')

    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_data/MA/sub0' + str(a) + '_test_data_04.csv',arr5, delimiter=',')
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_label/MA/sub0' + str(a) + '_test_label_04.csv',label5, delimiter='')

    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_data/MA/sub0' + str(a) + '_test_data_05.csv',arr6, delimiter=',')
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_label/MA/sub0' + str(a) + '_test_label_05.csv',label6, delimiter='')

    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_data/MA/sub0' + str(a) + '_test_data_06.csv',arr7, delimiter=',')
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_label/MA/sub0' + str(a) + '_test_label_06.csv',label7, delimiter='')

    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_data/MA/sub0' + str(a) + '_test_data_07.csv',arr8, delimiter=',')
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_label/MA/sub0' + str(a) + '_test_label_07.csv',label8, delimiter='')

    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_data/MA/sub0' + str(a) + '_test_data_08.csv',arr9, delimiter=',')
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_label/MA/sub0' + str(a) + '_test_label_08.csv',label9, delimiter='')

    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_data/MA/sub0' + str(a) + '_test_data_09.csv',arr10, delimiter=',')
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/EEG_data/testing_label/MA/sub0' + str(a) + '_test_label_09.csv',label10, delimiter='')