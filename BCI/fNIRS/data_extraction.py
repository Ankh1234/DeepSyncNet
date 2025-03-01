import numpy as np
import openpyxl
import sys
from scipy.interpolate import griddata
import matplotlib.pyplot as plt


def extract1(file, index=0):
    # 加载工作簿
    workbook = openpyxl.load_workbook(file, data_only=True)

    # 获取指定索引的工作表
    sheetnames = workbook.sheetnames
    worksheet = workbook[sheetnames[index]]

    # 获取所有行的数据
    all_data = []
    for row in worksheet.iter_rows(values_only=True):
        if any(val is not None for val in row):  # 跳过全空行
            all_data.append(row)

    # 转换为 NumPy 数组
    cc = np.array(all_data)
    return cc


for a in range(1,30):

    for b in range(2, 7, 2):
        data_session_extracted = []
        data_path = r'/root/autodl-tmp/project/Open_Access_BCI_Data/fNIRS_data/raw_data/MA/sub0'+str(a)+'_session0' + str(
            b) + '_data.xlsx'
        events_start_path = r'/root/autodl-tmp/project/Open_Access_BCI_Data/fNIRS_data/label_Hz/MA/sub0'+str(a)+'_session0' + str(
            b) + '_label_Hz.xlsx'
        label_path = r'/root/autodl-tmp/project/Open_Access_BCI_Data/fNIRS_data/label/MA/sub0'+str(a)+'_session0' + str(b) + '_label.xlsx'

        data = extract1(data_path)
        label = extract1(label_path)
        # data = data.T
        events = extract1(events_start_path)
        events = events.astype(int)
        label = label.astype(int)

        arr_3D_session = []
        label_s = []
        num_event = 0

        for i in events[0]:
            arr_time_window = []
            if label[num_event] == 1:
                num_event += 1
                for l in range(-2, 8):
                    arr = data[i + l * 10:i + l * 10 + 30]
                    arr_unit = []
                    for c in range(30):
                        arr_3D = np.full((16, 16), np.nan)

                        # 定义插值范围
                        interpolation_ranges = {
                            0: [6, 11 - 1],
                            1: [5, 12 - 1],
                            2: [4, 13 - 1],
                            3: [2, 15 - 1],
                            4: [2, 15 - 1],
                            5: [2, 15 - 1],
                            6: [2, 15 - 1],
                            7: [1, 16 - 1],
                            8: [2, 15 - 1],
                            9: [2, 15 - 1],
                            10: [2, 15 - 1],
                            11: [2, 15 - 1],
                            12: [4, 13 - 1],
                            13: [6, 11 - 1],
                            14: [6, 11 - 1],
                            15: [7, 10 - 1]
                        }

                        # 遍历矩阵行，设置待插值区域为 0，其他区域保持为 NaN
                        for row in range(16):
                            if row in interpolation_ranges:
                                start, end = interpolation_ranges[row]
                                arr_3D[row, start:end + 1] = 0  # 将待插值区域设置为 0
                            else:
                                arr_3D[row, :] = np.nan  # 非插值区域保持为 NaN

                        # 插入通道数据

                        # AF7~Fp1
                        arr_3D[3 - 1, 5 - 1] = arr[c, 0]
                        arr_3D[2 - 1, 6 - 1] = arr[c, 0]

                        # AF3~Fp1
                        arr_3D[3 - 1, 7 - 1] = (arr[c, 1] + arr[c, 2]) / 2  # AF3
                        arr_3D[2 - 1, 7 - 1] = arr[c, 1]

                        # AF3~AFz
                        arr_3D[3 - 1, 8 - 1] = arr[c, 2]

                        # Fpz~Fp1
                        arr_3D[1 - 1, 8 - 1] = arr[c, 3]
                        arr_3D[1 - 1, 7 - 1] = (arr[c, 0] + arr[c, 1] + arr[c, 3]) / 3  # Fp1

                        # Fpz~AFz
                        arr_3D[2 - 1, 9 - 1] = arr[c, 4]

                        # Fpz~Fp2
                        arr_3D[1 - 1, 10 - 1] = arr[c, 5]
                        arr_3D[1 - 1, 9 - 1] = (arr[c, 3] + arr[c, 4] + arr[c, 5]) / 3  # Fpz

                        # AF4~AFz
                        arr_3D[3 - 1, 10 - 1] = arr[c, 6]
                        arr_3D[3 - 1, 9 - 1] = (arr[c, 2] + arr[c, 4] + arr[c, 6]) / 3  # AFz

                        # AF4~Fp2
                        arr_3D[2 - 1, 11 - 1] = arr[c, 7]
                        arr_3D[3 - 1, 11 - 1] = (arr[c, 6] + arr[c, 7]) / 2

                        # AF8~Fp2
                        arr_3D[2 - 1, 12 - 1] = arr[c, 8]
                        arr_3D[3 - 1, 13 - 1] = arr[c, 8]
                        arr_3D[1 - 1, 11 - 1] = (arr[c, 5] + arr[c, 7] + arr[c, 8]) / 3

                        # Oz~POz
                        arr_3D[14 - 1, 9 - 1] = arr[c, 9]
                        arr_3D[15 - 1, 9 - 1] = arr[c, 9]

                        # Oz~O1
                        arr_3D[15 - 1, 7 - 1] = arr[c, 10]
                        arr_3D[15 - 1, 8 - 1] = arr[c, 10]
                        arr_3D[16 - 1, 8 - 1] = arr[c, 10]

                        # Oz~O2
                        arr_3D[15 - 1, 11 - 1] = arr[c, 11]
                        arr_3D[15 - 1, 10 - 1] = arr[c, 11]
                        arr_3D[16 - 1, 10 - 1] = arr[c, 11]
                        arr_3D[16 - 1, 9 - 1] = (arr[c, 9] + arr[c, 10] + arr[c, 11]) / 3  # Oz

                        # C5~CP5
                        arr_3D[9 - 1, 3 - 1] = arr[c, 12]

                        # C5~FC5
                        arr_3D[7 - 1, 3 - 1] = arr[c, 13]

                        # C5~C3
                        arr_3D[8 - 1, 4 - 1] = arr[c, 14]
                        arr_3D[8 - 1, 3 - 1] = (arr[c, 12] + arr[c, 13] + arr[c, 14]) / 3  # C5

                        # FC3~FC5
                        arr_3D[6 - 1, 4 - 1] = arr[c, 15]
                        arr_3D[6 - 1, 3 - 1] = (arr[c, 13] + arr[c, 15]) / 2  # FC5

                        # FC3~C3
                        arr_3D[7 - 1, 5 - 1] = arr[c, 16]

                        # FC3~FC1
                        arr_3D[6 - 1, 6 - 1] = arr[c, 17]
                        arr_3D[6 - 1, 5 - 1] = (arr[c, 15] + arr[c, 16] + arr[c, 17]) / 3  # FC3

                        # CP3~CP5
                        arr_3D[10 - 1, 4 - 1] = arr[c, 18]
                        arr_3D[10 - 1, 3 - 1] = (arr[c, 12] + arr[c, 18]) / 2  # CP5

                        # CP3~C3
                        arr_3D[9 - 1, 5 - 1] = arr[c, 19]

                        # CP3~CP1
                        arr_3D[10 - 1, 6 - 1] = arr[c, 20]
                        arr_3D[10 - 1, 5 - 1] = (arr[c, 18] + arr[c, 19] + arr[c, 20]) / 3  # CP3

                        # C1~C3
                        arr_3D[8 - 1, 6 - 1] = arr[c, 21]
                        arr_3D[8 - 1, 5 - 1] = (arr[c, 14] + arr[c, 16] + arr[c, 19] + arr[c, 21]) / 4  # C3

                        # C1~FC1
                        arr_3D[7 - 1, 7 - 1] = arr[c, 22]
                        arr_3D[6 - 1, 7 - 1] = (arr[c, 17] + arr[c, 22]) / 2  # FC1

                        # C1~CP1
                        arr_3D[9 - 1, 7 - 1] = arr[c, 23]
                        arr_3D[8 - 1, 7 - 1] = (arr[c, 21] + arr[c, 22] + arr[c, 23]) / 3  # C1
                        arr_3D[10 - 1, 7 - 1] = (arr[c, 20] + arr[c, 23]) / 2  # CP1

                        # 右侧
                        # C2~FC2
                        arr_3D[7 - 1, 11 - 1] = arr[c, 24]

                        # C2~CP2
                        arr_3D[9 - 1, 11 - 1] = arr[c, 25]

                        # C2~C4
                        arr_3D[8 - 1, 12 - 1] = arr[c, 26]
                        arr_3D[8 - 1, 11 - 1] = (arr[c, 24] + arr[c, 25] + arr[c, 26]) / 3  # C2

                        # FC4~FC2
                        arr_3D[6 - 1, 12 - 1] = arr[c, 27]
                        arr_3D[6 - 1, 11 - 1] = (arr[c, 24] + arr[c, 27]) / 2  # FC2

                        # FC4~C4
                        arr_3D[7 - 1, 13 - 1] = arr[c, 28]

                        # FC4~FC6
                        arr_3D[6 - 1, 14 - 1] = arr[c, 29]
                        arr_3D[6 - 1, 13 - 1] = (arr[c, 27] + arr[c, 28] + arr[c, 29]) / 3  # FC4

                        # CP4~CP6
                        arr_3D[10 - 1, 14 - 1] = arr[c, 30]

                        # CP4~CP2
                        arr_3D[10 - 1, 12 - 1] = arr[c, 31]
                        arr_3D[10 - 1, 11 - 1] = (arr[c, 25] + arr[c, 31]) / 2  # CP2

                        # CP4~C4
                        arr_3D[9 - 1, 13 - 1] = arr[c, 32]
                        arr_3D[10 - 1, 13 - 1] = (arr[c, 30] + arr[c, 31] + arr[c, 32]) / 3  # CP4

                        # C6~CP6
                        arr_3D[9 - 1, 15 - 1] = arr[c, 33]
                        arr_3D[10 - 1, 15 - 1] = (arr[c, 30] + arr[c, 33]) / 2  # CP6

                        # C6~C4
                        arr_3D[8 - 1, 14 - 1] = arr[c, 34]
                        arr_3D[8 - 1, 13 - 1] = (arr[c, 26] + arr[c, 28] + arr[c, 32] + arr[c, 34]) / 4  # C4

                        # C6~FC6
                        arr_3D[7 - 1, 15 - 1] = arr[c, 35]
                        arr_3D[6 - 1, 15 - 1] = (arr[c, 29] + arr[c, 35]) / 2  # FC6
                        arr_3D[8 - 1, 15 - 1] = (arr[c, 33] + arr[c, 34] + arr[c, 35]) / 3  # C6

                        # 获取已知点的坐标和值（通过遍历矩阵）
                        known_points = []
                        known_values = []

                        for row in range(16):
                            for col in range(16):
                                if not np.isnan(arr_3D[row, col]) and arr_3D[row, col] != 0:  # 确认值是已知点（非 NaN 且非 0）
                                    known_points.append((row, col))
                                    known_values.append(arr_3D[row, col])

                        # 获取插值区域中的目标点
                        target_points = []
                        for row in range(16):
                            if row in interpolation_ranges:
                                start, end = interpolation_ranges[row]
                                for col in range(start, end + 1):
                                    # 只添加待插值的点（值为 0 的点）
                                    if arr_3D[row, col] == 0:
                                        target_points.append((row, col))

                        # 使用 linear 方法对插值区域进行插值
                        interpolated_values_linear = griddata(
                            points=known_points,  # 已知点坐标
                            values=known_values,  # 已知点的值
                            xi=target_points,  # 插值目标点
                            method='linear'  # 线性插值
                        )

                        # 使用 nearest 方法对 NaN 边界点进行补充插值
                        interpolated_values_nearest = griddata(
                            points=known_points,  # 已知点坐标
                            values=known_values,  # 已知点的值
                            xi=target_points,  # 插值目标点
                            method='nearest'  # 最近邻插值
                        )

                        # 合并两种插值结果（优先使用 linear 的结果）
                        final_interpolated_values = []
                        for d, val in enumerate(interpolated_values_linear):
                            if np.isnan(val):  # 如果 linear 插值为 NaN，用 nearest 的值补充
                                final_interpolated_values.append(interpolated_values_nearest[d])
                            else:
                                final_interpolated_values.append(val)

                        # 将插值结果覆盖到矩阵中
                        for idx, (row, col) in enumerate(target_points):
                            arr_3D[row, col] = final_interpolated_values[idx]

                        for row in range(16):
                            for col in range(16):
                                if np.isnan(arr_3D[row, col]):  #将空白部分填充为0
                                    arr_3D[row, col] = 0


                        # 可视化插值结果
                        # plt.figure(figsize=(8, 8))
                        # plt.title("Interpolated Matrix")
                        # plt.imshow(arr_3D, cmap='viridis', origin='upper')
                        # plt.colorbar(label='Interpolated Values')
                        # plt.show()

                        arr_3D = arr_3D.tolist()
                        arr_unit.append(arr_3D)
                    # arr_unit = np.array(arr_unit)
                    # print(arr_unit.shape)
                    # arr_unit = arr_unit.tolist()
                    label_s.append(16)
                    arr_time_window.append(arr_unit)
            elif label[num_event] == 2:
                num_event += 1
                for l in range(-2, 8):
                    arr = data[i + l * 10:i + l * 10 + 30]
                    arr_unit = []
                    for c in range(30):
                        arr_3D = np.full((16, 16), np.nan)

                        # 定义插值范围
                        interpolation_ranges = {
                            0: [6, 11 - 1],
                            1: [5, 12 - 1],
                            2: [4, 13 - 1],
                            3: [2, 15 - 1],
                            4: [2, 15 - 1],
                            5: [2, 15 - 1],
                            6: [2, 15 - 1],
                            7: [1, 16 - 1],
                            8: [2, 15 - 1],
                            9: [2, 15 - 1],
                            10: [2, 15 - 1],
                            11: [2, 15 - 1],
                            12: [4, 13 - 1],
                            13: [6, 11 - 1],
                            14: [6, 11 - 1],
                            15: [7, 10 - 1]
                        }

                        # 遍历矩阵行，设置待插值区域为 0，其他区域保持为 NaN
                        for row in range(16):
                            if row in interpolation_ranges:
                                start, end = interpolation_ranges[row]
                                arr_3D[row, start:end + 1] = 0  # 将待插值区域设置为 0
                            else:
                                arr_3D[row, :] = np.nan  # 非插值区域保持为 NaN

                        # 插入通道数据

                        # AF7~Fp1
                        arr_3D[3 - 1, 5 - 1] = arr[c, 0]
                        arr_3D[2 - 1, 6 - 1] = arr[c, 0]

                        # AF3~Fp1
                        arr_3D[3 - 1, 7 - 1] = (arr[c, 1] + arr[c, 2]) / 2  # AF3
                        arr_3D[2 - 1, 7 - 1] = arr[c, 1]

                        # AF3~AFz
                        arr_3D[3 - 1, 8 - 1] = arr[c, 2]

                        # Fpz~Fp1
                        arr_3D[1 - 1, 8 - 1] = arr[c, 3]
                        arr_3D[1 - 1, 7 - 1] = (arr[c, 0] + arr[c, 1] + arr[c, 3]) / 3  # Fp1

                        # Fpz~AFz
                        arr_3D[2 - 1, 9 - 1] = arr[c, 4]

                        # Fpz~Fp2
                        arr_3D[1 - 1, 10 - 1] = arr[c, 5]
                        arr_3D[1 - 1, 9 - 1] = (arr[c, 3] + arr[c, 4] + arr[c, 5]) / 3  # Fpz

                        # AF4~AFz
                        arr_3D[3 - 1, 10 - 1] = arr[c, 6]
                        arr_3D[3 - 1, 9 - 1] = (arr[c, 2] + arr[c, 4] + arr[c, 6]) / 3  # AFz

                        # AF4~Fp2
                        arr_3D[2 - 1, 11 - 1] = arr[c, 7]
                        arr_3D[3 - 1, 11 - 1] = (arr[c, 6] + arr[c, 7]) / 2

                        # AF8~Fp2
                        arr_3D[2 - 1, 12 - 1] = arr[c, 8]
                        arr_3D[3 - 1, 13 - 1] = arr[c, 8]
                        arr_3D[1 - 1, 11 - 1] = (arr[c, 5] + arr[c, 7] + arr[c, 8]) / 3

                        # Oz~POz
                        arr_3D[14 - 1, 9 - 1] = arr[c, 9]
                        arr_3D[15 - 1, 9 - 1] = arr[c, 9]

                        # Oz~O1
                        arr_3D[15 - 1, 7 - 1] = arr[c, 10]
                        arr_3D[15 - 1, 8 - 1] = arr[c, 10]
                        arr_3D[16 - 1, 8 - 1] = arr[c, 10]

                        # Oz~O2
                        arr_3D[15 - 1, 11 - 1] = arr[c, 11]
                        arr_3D[15 - 1, 10 - 1] = arr[c, 11]
                        arr_3D[16 - 1, 10 - 1] = arr[c, 11]
                        arr_3D[16 - 1, 9 - 1] = (arr[c, 9] + arr[c, 10] + arr[c, 11]) / 3  # Oz

                        # C5~CP5
                        arr_3D[9 - 1, 3 - 1] = arr[c, 12]

                        # C5~FC5
                        arr_3D[7 - 1, 3 - 1] = arr[c, 13]

                        # C5~C3
                        arr_3D[8 - 1, 4 - 1] = arr[c, 14]
                        arr_3D[8 - 1, 3 - 1] = (arr[c, 12] + arr[c, 13] + arr[c, 14]) / 3  # C5

                        # FC3~FC5
                        arr_3D[6 - 1, 4 - 1] = arr[c, 15]
                        arr_3D[6 - 1, 3 - 1] = (arr[c, 13] + arr[c, 15]) / 2  # FC5

                        # FC3~C3
                        arr_3D[7 - 1, 5 - 1] = arr[c, 16]

                        # FC3~FC1
                        arr_3D[6 - 1, 6 - 1] = arr[c, 17]
                        arr_3D[6 - 1, 5 - 1] = (arr[c, 15] + arr[c, 16] + arr[c, 17]) / 3  # FC3

                        # CP3~CP5
                        arr_3D[10 - 1, 4 - 1] = arr[c, 18]
                        arr_3D[10 - 1, 3 - 1] = (arr[c, 12] + arr[c, 18]) / 2  # CP5

                        # CP3~C3
                        arr_3D[9 - 1, 5 - 1] = arr[c, 19]

                        # CP3~CP1
                        arr_3D[10 - 1, 6 - 1] = arr[c, 20]
                        arr_3D[10 - 1, 5 - 1] = (arr[c, 18] + arr[c, 19] + arr[c, 20]) / 3  # CP3

                        # C1~C3
                        arr_3D[8 - 1, 6 - 1] = arr[c, 21]
                        arr_3D[8 - 1, 5 - 1] = (arr[c, 14] + arr[c, 16] + arr[c, 19] + arr[c, 21]) / 4  # C3

                        # C1~FC1
                        arr_3D[7 - 1, 7 - 1] = arr[c, 22]
                        arr_3D[6 - 1, 7 - 1] = (arr[c, 17] + arr[c, 22]) / 2  # FC1

                        # C1~CP1
                        arr_3D[9 - 1, 7 - 1] = arr[c, 23]
                        arr_3D[8 - 1, 7 - 1] = (arr[c, 21] + arr[c, 22] + arr[c, 23]) / 3  # C1
                        arr_3D[10 - 1, 7 - 1] = (arr[c, 20] + arr[c, 23]) / 2  # CP1

                        # 右侧
                        # C2~FC2
                        arr_3D[7 - 1, 11 - 1] = arr[c, 24]

                        # C2~CP2
                        arr_3D[9 - 1, 11 - 1] = arr[c, 25]

                        # C2~C4
                        arr_3D[8 - 1, 12 - 1] = arr[c, 26]
                        arr_3D[8 - 1, 11 - 1] = (arr[c, 24] + arr[c, 25] + arr[c, 26]) / 3  # C2

                        # FC4~FC2
                        arr_3D[6 - 1, 12 - 1] = arr[c, 27]
                        arr_3D[6 - 1, 11 - 1] = (arr[c, 24] + arr[c, 27]) / 2  # FC2

                        # FC4~C4
                        arr_3D[7 - 1, 13 - 1] = arr[c, 28]

                        # FC4~FC6
                        arr_3D[6 - 1, 14 - 1] = arr[c, 29]
                        arr_3D[6 - 1, 13 - 1] = (arr[c, 27] + arr[c, 28] + arr[c, 29]) / 3  # FC4

                        # CP4~CP6
                        arr_3D[10 - 1, 14 - 1] = arr[c, 30]

                        # CP4~CP2
                        arr_3D[10 - 1, 12 - 1] = arr[c, 31]
                        arr_3D[10 - 1, 11 - 1] = (arr[c, 25] + arr[c, 31]) / 2  # CP2

                        # CP4~C4
                        arr_3D[9 - 1, 13 - 1] = arr[c, 32]
                        arr_3D[10 - 1, 13 - 1] = (arr[c, 30] + arr[c, 31] + arr[c, 32]) / 3  # CP4

                        # C6~CP6
                        arr_3D[9 - 1, 15 - 1] = arr[c, 33]
                        arr_3D[10 - 1, 15 - 1] = (arr[c, 30] + arr[c, 33]) / 2  # CP6

                        # C6~C4
                        arr_3D[8 - 1, 14 - 1] = arr[c, 34]
                        arr_3D[8 - 1, 13 - 1] = (arr[c, 26] + arr[c, 28] + arr[c, 32] + arr[c, 34]) / 4  # C4

                        # C6~FC6
                        arr_3D[7 - 1, 15 - 1] = arr[c, 35]
                        arr_3D[6 - 1, 15 - 1] = (arr[c, 29] + arr[c, 35]) / 2  # FC6
                        arr_3D[8 - 1, 15 - 1] = (arr[c, 33] + arr[c, 34] + arr[c, 35]) / 3  # C6

                        # 获取已知点的坐标和值（通过遍历矩阵）
                        known_points = []
                        known_values = []

                        for row in range(16):
                            for col in range(16):
                                if not np.isnan(arr_3D[row, col]) and arr_3D[row, col] != 0:  # 确认值是已知点（非 NaN 且非 0）
                                    known_points.append((row, col))
                                    known_values.append(arr_3D[row, col])

                        # 获取插值区域中的目标点
                        target_points = []
                        for row in range(16):
                            if row in interpolation_ranges:
                                start, end = interpolation_ranges[row]
                                for col in range(start, end + 1):
                                    # 只添加待插值的点（值为 0 的点）
                                    if arr_3D[row, col] == 0:
                                        target_points.append((row, col))

                        # 使用 linear 方法对插值区域进行插值
                        interpolated_values_linear = griddata(
                            points=known_points,  # 已知点坐标
                            values=known_values,  # 已知点的值
                            xi=target_points,  # 插值目标点
                            method='cubic'  # 三次插值
                        )

                        # 使用 nearest 方法对 NaN 边界点进行补充插值
                        interpolated_values_nearest = griddata(
                            points=known_points,  # 已知点坐标
                            values=known_values,  # 已知点的值
                            xi=target_points,  # 插值目标点
                            method='nearest'  # 最近邻插值
                        )

                        # 合并两种插值结果（优先使用 linear 的结果）
                        final_interpolated_values = []
                        for d, val in enumerate(interpolated_values_linear):
                            if np.isnan(val):  # 如果 linear 插值为 NaN，用 nearest 的值补充
                                final_interpolated_values.append(interpolated_values_nearest[d])
                            else:
                                final_interpolated_values.append(val)

                        # 将插值结果覆盖到矩阵中
                        for idx, (row, col) in enumerate(target_points):
                            arr_3D[row, col] = final_interpolated_values[idx]

                        for row in range(16):
                            for col in range(16):
                                if np.isnan(arr_3D[row, col]):  #将空白部分填充为0
                                    arr_3D[row, col] = 0


                        # 可视化插值结果
                        #plt.figure(figsize=(8, 8))
                        #plt.title("Interpolated Matrix")
                        #plt.imshow(arr_3D, cmap='viridis', origin='upper')
                        #plt.colorbar(label='Interpolated Values')
                        #plt.show()

                        arr_3D = arr_3D.tolist()
                        arr_unit.append(arr_3D)
                    label_s.append(32)
                    arr_time_window.append(arr_unit)
            else:
                # print("Wrong")
                sys.exit()

            arr_3D_session.append(arr_time_window)
        data_extracted = np.array(arr_3D_session)
        print(data_extracted.shape)
        data_extracted = data_extracted.reshape(200, -1)
        print(data_extracted.shape)
        if b == 2:
            data_session2_extracted = data_extracted
            label_session2_extracted = label_s
        elif b == 4:
            data_session4_extracted = data_extracted
            label_session4_extracted = label_s
        else:
            data_session6_extracted = data_extracted
            label_session6_extracted = label_s
    data_sub_extracted = np.concatenate((data_session2_extracted, data_session4_extracted, data_session6_extracted),
                                        axis=0)
    label_sub_extracted = np.concatenate((label_session2_extracted, label_session4_extracted, label_session6_extracted),
                                         axis=0)
    print(data_sub_extracted.shape)
    print(label_sub_extracted.shape)
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/fNIRS_data/extracted_data/MA/sub0'+str(a)+'_extracted_data.csv', data_sub_extracted, delimiter=',')
    np.savetxt('/root/autodl-tmp/project/Open_Access_BCI_Data/fNIRS_data/extracted_label/MA/sub0' + str(a) + '_extracted_label.csv', label_sub_extracted, delimiter='')


"""

for i in events:
    arr = data[i:i+2000]
    arr = arr.T
    data_extracted = np.append(data_extracted,arr,axis=0)
print(data_extracted.shape)

df_data = pd.read_excel(data_path, sheet_name = "Sheet1") # sheet_name不指定时默认返回全表数据
df_events = pd.read_excel(events_start_path, sheet_name = "Sheet1")
df_label = pd.read_excel(label_path, sheet_name = "Sheet1")

events = list(df_events.columns)
print(events)
data = df_data.iloc[events[0]:events[0]+2000]
data = data.T
print(data.shape)
data.to_csv("/Users/lihao/Desktop/Open_Access_BCI_Data/EEG_data/extracted_data/test1.csv")


for a in range(1,20):
    single_event_data = df_data.iloc[events[a]:events[a]+2000]
    single_event_data = single_event_data.T
    data = data.append(single_event_data)


#打印5行看看数据对不对
data.to_csv("/Users/lihao/Desktop/Open_Access_BCI_Data/EEG_data/extracted_data/test2.csv")
print(data.shape)

"""